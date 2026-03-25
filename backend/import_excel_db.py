"""
import_excel_db.py – Import HMS_Data_Base.xlsx into the Neon PostgreSQL database.

Reads all sheets (Patient, Doctor, Nurse, Receptionist, Lab Technician, Pharmacist)
and inserts/updates records in the users, staff, and patients tables.

Usage (from the backend/ directory):
    python import_excel_db.py

Or from the project root:
    python backend/import_excel_db.py
"""

import os
import sys
from pathlib import Path

import psycopg2
import psycopg2.extras
from psycopg2 import sql
import openpyxl
from dotenv import load_dotenv

# ── Resolve paths & load .env ─────────────────────────────────────────────────
_script_dir   = Path(__file__).resolve().parent
_project_root = _script_dir.parent

load_dotenv(_script_dir / ".env")
load_dotenv(_project_root / ".env", override=False)

sys.path.insert(0, str(_script_dir))
from app.auth import hash_password  # noqa: E402

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL_PATH     = _project_root / "HMS_Data_Base.xlsx"
DEFAULT_PW     = "Medcore@123"

ROLE_COLORS = {
    "Doctor":         "3B82F6",
    "Nurse":          "10B981",
    "Receptionist":   "F59E0B",
    "Lab Technician": "8B5CF6",
    "Pharmacist":     "EC4899",
    "Patient":        "6B7280",
}

KNOWN_SHEETS = {
    "Doctor",
    "Nurse",
    "Receptionist",
    "Lab Technician",
    "Pharmacist",
    "Patient",
}

HEADER_ALIASES = {
    "name": "full_name",
    "patientname": "full_name",
    "doctorname": "full_name",
    "nursename": "full_name",
    "fullname": "full_name",
    "emailid": "email",
    "phoneno": "contact",
    "phonenumber": "contact",
    "phone": "contact",
    "mobileno": "contact",
    "bednumber": "bed_number",
    "wardname": "ward",
    "heartrate": "heart_rate",
    "bloodpressure": "bp",
    "oxygen": "spo2",
    "oxygenlevel": "spo2",
    "lastupdated": "last_updated",
    "instruction": "instructions",
}


def avatar(name: str, role: str) -> str:
    color  = ROLE_COLORS.get(role, "6B7280")
    encoded = "+".join(name.replace("Dr. ", "").split())
    return f"https://ui-avatars.com/api/?name={encoded}&background={color}&color=fff"


def normalise_shift(raw: str | None) -> str:
    """Map 'Morning Shift' / 'Morning' etc. to the DB-allowed values."""
    if not raw:
        return "Morning"
    r = str(raw).strip().lower()
    if "morning" in r:
        return "Morning"
    if "evening" in r:
        return "Evening"
    if "night" in r:
        return "Night"
    return "Morning"


def load_sheet(wb: openpyxl.Workbook, sheet_name: str) -> list[dict]:
    """
    Each sheet has one blank row, then a header row, then data rows.
    Returns a list of dicts keyed by header column names.
    """
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Find the header row (first row that has at least 2 non-None cells)
    header_idx = None
    for i, row in enumerate(rows):
        non_null = [c for c in row if c is not None]
        if len(non_null) >= 2:
            header_idx = i
            break

    if header_idx is None:
        return []

    headers = [str(h).strip() if h is not None else f"col_{i}"
               for i, h in enumerate(rows[header_idx])]

    records = []
    for row in rows[header_idx + 1:]:
        if not any(v is not None for v in row):
            continue
        records.append(dict(zip(headers, row)))
    return records


def normalise_key(text: str | None) -> str:
    if text is None:
        return ""
    return "".join(ch.lower() for ch in str(text).strip() if ch.isalnum())


def normalise_column(text: str | None) -> str:
    if text is None:
        return ""
    stripped = "".join(ch.lower() if ch.isalnum() else "_" for ch in str(text).strip())
    while "__" in stripped:
        stripped = stripped.replace("__", "_")
    return stripped.strip("_")


def get_public_table_columns(cur) -> dict[str, list[dict]]:
    cur.execute(
        """
        SELECT
            table_name,
            column_name,
            data_type,
            is_nullable,
            column_default,
            is_identity
        FROM information_schema.columns
        WHERE table_schema = 'public'
        ORDER BY table_name, ordinal_position
        """
    )
    rows = cur.fetchall()
    out: dict[str, list[dict]] = {}
    for table_name, column_name, data_type, is_nullable, column_default, is_identity in rows:
        out.setdefault(table_name, []).append({
            "column_name": column_name,
            "data_type": data_type,
            "is_nullable": is_nullable,
            "column_default": column_default,
            "is_identity": is_identity,
        })
    return out


def _tokenise_name(text: str) -> set[str]:
    tokens = []
    current = []
    for ch in text:
        if ch.isalnum():
            current.append(ch.lower())
        else:
            if current:
                tokens.append("".join(current))
                current = []
    if current:
        tokens.append("".join(current))
    return set(tokens)


def find_target_table(sheet_name: str, records: list[dict], tables: dict[str, list[dict]]) -> str | None:
    table_names = list(tables.keys())
    direct = {normalise_key(t): t for t in table_names}
    sheet_norm = normalise_key(sheet_name)

    if sheet_norm in direct:
        return direct[sheet_norm]
    if sheet_norm.endswith("s") and sheet_norm[:-1] in direct:
        return direct[sheet_norm[:-1]]
    if (sheet_norm + "s") in direct:
        return direct[sheet_norm + "s"]

    headers = list(records[0].keys()) if records else []
    header_norms = [normalise_key(h) for h in headers]
    header_targets = {
        normalise_key(HEADER_ALIASES.get(hn, normalise_column(hn)))
        for hn in header_norms
    }

    best_table = None
    best_score = 0
    sheet_tokens = _tokenise_name(sheet_name)

    for table_name, cols in tables.items():
        table_norm = normalise_key(table_name)
        table_tokens = _tokenise_name(table_name)
        col_norms = {normalise_key(c["column_name"]) for c in cols}

        score = 0
        score += 3 * len(sheet_tokens & table_tokens)
        score += 5 * len(header_targets & col_norms)

        if sheet_norm == table_norm:
            score += 100
        elif sheet_norm.endswith("s") and sheet_norm[:-1] == table_norm:
            score += 95
        elif sheet_norm + "s" == table_norm:
            score += 95

        if score > best_score:
            best_score = score
            best_table = table_name

    return best_table if best_score >= 8 else None


def convert_value(value, data_type: str):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None

    numeric_types = {
        "smallint",
        "integer",
        "bigint",
        "numeric",
        "real",
        "double precision",
        "decimal",
    }
    if data_type in numeric_types:
        if isinstance(value, (int, float)):
            return value
        try:
            text = str(value).replace(",", "")
            return float(text) if "." in text else int(text)
        except Exception:
            return None

    if data_type == "boolean":
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in {"true", "t", "yes", "y", "1", "active"}:
            return True
        if text in {"false", "f", "no", "n", "0", "inactive"}:
            return False
        return None

    return value


def insert_dynamic_sheet(cur, sheet_name: str, records: list[dict], tables: dict[str, list[dict]]) -> tuple[int, int]:
    table_name = find_target_table(sheet_name, records, tables)
    if table_name is None:
        print(f"↷ Sheet '{sheet_name}' skipped (no matching public table)")
        return 0, len(records)

    column_meta = tables[table_name]
    columns_by_norm: dict[str, str] = {}
    data_type_by_column: dict[str, str] = {}
    required_columns: set[str] = set()
    identity_columns: set[str] = set()

    for col in column_meta:
        col_name = col["column_name"]
        columns_by_norm[normalise_key(col_name)] = col_name
        data_type_by_column[col_name] = col["data_type"]
        if col["is_identity"] == "YES" or col_name == "id":
            identity_columns.add(col_name)

        is_required = (
            col["is_nullable"] == "NO"
            and col["column_default"] is None
            and col["is_identity"] != "YES"
            and col_name != "id"
        )
        if is_required:
            required_columns.add(col_name)

    inserted = 0
    skipped = 0

    for row in records:
        mapped: dict[str, object] = {}
        for raw_header, raw_value in row.items():
            header_norm = normalise_key(raw_header)
            alias_column = HEADER_ALIASES.get(header_norm)

            candidates = [
                alias_column,
                columns_by_norm.get(header_norm),
                columns_by_norm.get(normalise_key(normalise_column(raw_header))),
            ]

            target_column = next((c for c in candidates if c and c in data_type_by_column), None)
            if not target_column:
                continue
            if target_column in identity_columns:
                continue

            mapped[target_column] = convert_value(raw_value, data_type_by_column[target_column])

        if table_name == "nurse_orders":
            # Ensure patient_id is valid; fallback to patient_name lookup from sheet.
            patient_id = mapped.get("patient_id")
            patient_name_raw = row.get("patient_name")
            patient_name = str(patient_name_raw).strip() if patient_name_raw is not None else ""

            if patient_id is not None:
                cur.execute("SELECT 1 FROM patients WHERE id = %s", (patient_id,))
                if cur.fetchone() is None:
                    mapped["patient_id"] = None

            if mapped.get("patient_id") is None and patient_name:
                cur.execute(
                    """
                    SELECT id
                    FROM patients
                    WHERE lower(full_name) = lower(%s)
                    ORDER BY id
                    LIMIT 1
                    """,
                    (patient_name,),
                )
                patient_match = cur.fetchone()
                if patient_match:
                    mapped["patient_id"] = patient_match[0]

            if mapped.get("doctor_id") is None:
                patient_id = mapped.get("patient_id")
                if patient_id is not None:
                    cur.execute(
                        """
                        SELECT doctor_id
                        FROM appointments
                        WHERE patient_id = %s
                        ORDER BY appointment_date DESC NULLS LAST, id DESC
                        LIMIT 1
                        """,
                        (patient_id,),
                    )
                    appt_doctor = cur.fetchone()
                    if appt_doctor:
                        mapped["doctor_id"] = appt_doctor[0]

            if mapped.get("doctor_id") is None:
                cur.execute(
                    """
                    SELECT id
                    FROM users
                    WHERE role = 'Doctor'
                    ORDER BY id
                    LIMIT 1
                    """
                )
                fallback_doctor = cur.fetchone()
                if fallback_doctor:
                    mapped["doctor_id"] = fallback_doctor[0]

        if not mapped:
            skipped += 1
            continue

        missing_required = [c for c in required_columns if mapped.get(c) is None]
        if missing_required:
            skipped += 1
            continue

        cols = list(mapped.keys())
        values = [mapped[c] for c in cols]
        query = sql.SQL("INSERT INTO {table} ({cols}) VALUES ({vals}) ON CONFLICT DO NOTHING").format(
            table=sql.Identifier(table_name),
            cols=sql.SQL(", ").join(sql.Identifier(c) for c in cols),
            vals=sql.SQL(", ").join(sql.Placeholder() for _ in cols),
        )
        cur.execute("SAVEPOINT dynamic_row")
        try:
            cur.execute(query, values)
            inserted += 1
            cur.execute("RELEASE SAVEPOINT dynamic_row")
        except psycopg2.Error:
            cur.execute("ROLLBACK TO SAVEPOINT dynamic_row")
            skipped += 1

    print(f"✓ Dynamic   – Sheet '{sheet_name}' -> '{table_name}': {inserted} inserted, {skipped} skipped")
    return inserted, skipped


# ── Main ──────────────────────────────────────────────────────────────────────

def run():
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL is not set in .env")

    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel file not found at: {EXCEL_PATH}")

    print(f"📂 Loading workbook: {EXCEL_PATH.name}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    default_hash = hash_password(DEFAULT_PW)

    conn = psycopg2.connect(database_url)
    cur  = conn.cursor()

    # ── DOCTORS ───────────────────────────────────────────────────────────────
    doctors = load_sheet(wb, "Doctor")
    doctor_rows = []
    for d in doctors:
        name  = str(d.get("Name", "")).strip()
        email = str(d.get("Email", "")).strip().lower()
        phone = str(d.get("Phone No", "") or "").strip()
        dept  = str(d.get("Department", "") or "").strip()
        shift = normalise_shift(d.get("Working Shift"))
        bio   = str(d.get("Professional Bio", "") or "").strip()
        if not name or not email:
            continue
        doctor_rows.append((
            name, email, default_hash, "Doctor",
            dept, phone, "Active", shift, avatar(name, "Doctor"), bio, None
        ))

    psycopg2.extras.execute_batch(
        cur,
        """
        INSERT INTO users
            (full_name, email, password_hash, role, department, contact,
             status, shift, avatar_url, bio, consultation_fee)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (email) DO UPDATE SET
            full_name        = EXCLUDED.full_name,
            role             = EXCLUDED.role,
            department       = EXCLUDED.department,
            contact          = EXCLUDED.contact,
            status           = EXCLUDED.status,
            shift            = EXCLUDED.shift,
            avatar_url       = EXCLUDED.avatar_url,
            bio              = EXCLUDED.bio
        """,
        doctor_rows,
    )

    # Staff table for doctors
    staff_doctor_rows = [
        (r[0], r[1], "Doctor", r[4], r[5], r[6], r[7], r[8])
        for r in doctor_rows
    ]
    psycopg2.extras.execute_batch(
        cur,
        """
        INSERT INTO staff
            (full_name, email, role, department, contact, status, shift, avatar_url)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (email) DO UPDATE SET
            full_name  = EXCLUDED.full_name,
            role       = EXCLUDED.role,
            department = EXCLUDED.department,
            contact    = EXCLUDED.contact,
            status     = EXCLUDED.status,
            shift      = EXCLUDED.shift,
            avatar_url = EXCLUDED.avatar_url
        """,
        staff_doctor_rows,
    )
    conn.commit()
    print(f"✓ Doctors   – {len(doctor_rows)} records inserted/updated")

    # ── NURSES ────────────────────────────────────────────────────────────────
    nurses = load_sheet(wb, "Nurse")
    nurse_rows = []
    for n in nurses:
        name  = str(n.get("Name", "")).strip()
        email = str(n.get("Email", "")).strip().lower()
        phone = str(n.get("Phone", "") or "").strip()
        dept  = str(n.get("Department", "") or "").strip()
        shift = normalise_shift(n.get("Shift"))
        bio   = str(n.get("Specialization / Notes", "") or "").strip()
        if not name or not email:
            continue
        nurse_rows.append((
            name, email, default_hash, "Nurse",
            dept, phone, "Active", shift, avatar(name, "Nurse"), bio, None
        ))

    psycopg2.extras.execute_batch(
        cur,
        """
        INSERT INTO users
            (full_name, email, password_hash, role, department, contact,
             status, shift, avatar_url, bio, consultation_fee)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (email) DO NOTHING
        """,
        nurse_rows,
    )

    staff_nurse_rows = [
        (r[0], r[1], "Nurse", r[4], r[5], r[6], r[7], r[8])
        for r in nurse_rows
    ]
    psycopg2.extras.execute_batch(
        cur,
        """
        INSERT INTO staff
            (full_name, email, role, department, contact, status, shift, avatar_url)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (email) DO NOTHING
        """,
        staff_nurse_rows,
    )
    conn.commit()
    print(f"✓ Nurses    – {len(nurse_rows)} records inserted/updated")

    # ── RECEPTIONISTS ─────────────────────────────────────────────────────────
    receptionists = load_sheet(wb, "Receptionist")
    recept_rows = []
    for r in receptionists:
        name  = str(r.get("Name", "")).strip()
        email = str(r.get("Email", "")).strip().lower()
        phone = str(r.get("Phone No", "") or "").strip()
        dept  = str(r.get("Department", "") or "").strip()
        shift = normalise_shift(r.get("Working Shift"))
        if not name or not email:
            continue
        recept_rows.append((
            name, email, default_hash, "Receptionist",
            dept, phone, "Active", shift, avatar(name, "Receptionist"), None, None
        ))

    psycopg2.extras.execute_batch(
        cur,
        """
        INSERT INTO users
            (full_name, email, password_hash, role, department, contact,
             status, shift, avatar_url, bio, consultation_fee)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (email) DO NOTHING
        """,
        recept_rows,
    )

    staff_recept_rows = [
        (r[0], r[1], "Receptionist", r[4], r[5], r[6], r[7], r[8])
        for r in recept_rows
    ]
    psycopg2.extras.execute_batch(
        cur,
        """
        INSERT INTO staff
            (full_name, email, role, department, contact, status, shift, avatar_url)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (email) DO NOTHING
        """,
        staff_recept_rows,
    )
    conn.commit()
    print(f"✓ Receptnst – {len(recept_rows)} records inserted/updated")

    # ── LAB TECHNICIANS ───────────────────────────────────────────────────────
    lab_techs = load_sheet(wb, "Lab Technician")
    lt_rows = []
    for lt in lab_techs:
        name  = str(lt.get("Name", "")).strip()
        email = str(lt.get("Email", "")).strip().lower()
        phone = str(lt.get("Phone No", "") or "").strip()
        dept  = str(lt.get("Department", "") or "").strip()
        shift = normalise_shift(lt.get("Working Shift"))
        bio   = str(lt.get("Certification / Notes", "") or "").strip()
        if not name or not email:
            continue
        lt_rows.append((
            name, email, default_hash, "Lab Technician",
            dept, phone, "Active", shift, avatar(name, "Lab Technician"), bio, None
        ))

    psycopg2.extras.execute_batch(
        cur,
        """
        INSERT INTO users
            (full_name, email, password_hash, role, department, contact,
             status, shift, avatar_url, bio, consultation_fee)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (email) DO NOTHING
        """,
        lt_rows,
    )

    staff_lt_rows = [
        (r[0], r[1], "Lab Technician", r[4], r[5], r[6], r[7], r[8])
        for r in lt_rows
    ]
    psycopg2.extras.execute_batch(
        cur,
        """
        INSERT INTO staff
            (full_name, email, role, department, contact, status, shift, avatar_url)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (email) DO NOTHING
        """,
        staff_lt_rows,
    )
    conn.commit()
    print(f"✓ Lab Techs – {len(lt_rows)} records inserted/updated")

    # ── PHARMACISTS ───────────────────────────────────────────────────────────
    pharmacists = load_sheet(wb, "Pharmacist")
    ph_rows = []
    for ph in pharmacists:
        name  = str(ph.get("Name", "")).strip()
        email = str(ph.get("Email", "")).strip().lower()
        phone = str(ph.get("Phone No", "") or "").strip()
        dept  = str(ph.get("Department", "") or "").strip()
        shift = normalise_shift(ph.get("Working Shift"))
        if not name or not email:
            continue
        ph_rows.append((
            name, email, default_hash, "Pharmacist",
            dept, phone, "Active", shift, avatar(name, "Pharmacist"), None, None
        ))

    psycopg2.extras.execute_batch(
        cur,
        """
        INSERT INTO users
            (full_name, email, password_hash, role, department, contact,
             status, shift, avatar_url, bio, consultation_fee)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (email) DO NOTHING
        """,
        ph_rows,
    )

    staff_ph_rows = [
        (r[0], r[1], "Pharmacist", r[4], r[5], r[6], r[7], r[8])
        for r in ph_rows
    ]
    psycopg2.extras.execute_batch(
        cur,
        """
        INSERT INTO staff
            (full_name, email, role, department, contact, status, shift, avatar_url)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (email) DO NOTHING
        """,
        staff_ph_rows,
    )
    conn.commit()
    print(f"✓ Pharmacst – {len(ph_rows)} records inserted/updated")

    # ── PATIENTS ──────────────────────────────────────────────────────────────
    patients = load_sheet(wb, "Patient")

    inserted_patients  = 0
    updated_patients   = 0
    linked_patients    = 0

    for p in patients:
        name  = str(p.get("Patient Name", "")).strip()
        email = str(p.get("Email ID", "")).strip().lower()
        if not name or not email:
            continue

        # Parse clinical fields
        age            = p.get("age")
        gender_raw     = str(p.get("gender", "") or "").strip()
        gender         = gender_raw if gender_raw in ("Male", "Female", "Other") else None

        # contact may arrive as a numeric value from Excel
        contact_raw = p.get("contact")
        if contact_raw is None:
            contact = None
        elif isinstance(contact_raw, (int, float)):
            contact = str(int(contact_raw))
        else:
            contact = str(contact_raw).strip() or None

        last_visit_raw    = p.get("last_visit")
        last_visit        = last_visit_raw.date() if hasattr(last_visit_raw, "date") else last_visit_raw
        medical_condition = str(p.get("medical_condition", "") or "").strip() or None
        pat_status_raw    = str(p.get("status", "") or "Outpatient").strip()
        pat_status        = pat_status_raw if pat_status_raw in ("Inpatient", "Outpatient", "Discharged") else "Outpatient"
        blood_type        = str(p.get("blood_type", "") or "").strip() or None
        allergies         = str(p.get("allergies", "") or "").strip() or None

        # Vitals
        bp_val      = str(p.get("bp", "") or "").strip() or None
        hr_raw      = p.get("heart_rate")
        heart_rate  = str(int(hr_raw)) if isinstance(hr_raw, (int, float)) else (str(hr_raw or "").strip() or None)
        temperature = str(p.get("temperature", "") or "").strip() or None
        spo2_raw    = p.get("spo2")
        spo2        = str(int(spo2_raw)) if isinstance(spo2_raw, (int, float)) else (str(spo2_raw or "").strip() or None)

        # 1. Get or create Patient user account
        cur.execute("SELECT id, role FROM users WHERE email = %s", (email,))
        existing = cur.fetchone()

        if existing is None:
            cur.execute(
                """
                INSERT INTO users
                    (full_name, email, password_hash, role, status, avatar_url)
                VALUES (%s, %s, %s, 'Patient', 'Active', %s)
                RETURNING id
                """,
                (name, email, default_hash, avatar(name, "Patient")),
            )
            user_id = cur.fetchone()[0]
            inserted_patients += 1
        else:
            user_id = existing[0]
            linked_patients += 1

        # 2. Upsert patients clinical record – update ALL fields from the sheet
        cur.execute("SELECT id FROM patients WHERE user_id = %s", (user_id,))
        existing_patient = cur.fetchone()

        if existing_patient is None:
            cur.execute(
                """
                INSERT INTO patients
                    (full_name, user_id, age, gender, contact, last_visit,
                     medical_condition, status, blood_type, allergies)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (name, user_id, age, gender, contact, last_visit,
                 medical_condition, pat_status, blood_type, allergies),
            )
            patient_id = cur.fetchone()[0]
        else:
            patient_id = existing_patient[0]
            cur.execute(
                """
                UPDATE patients SET
                    full_name         = %s,
                    age               = %s,
                    gender            = %s,
                    contact           = %s,
                    last_visit        = %s,
                    medical_condition = %s,
                    status            = %s,
                    blood_type        = %s,
                    allergies         = %s
                WHERE id = %s
                """,
                (name, age, gender, contact, last_visit,
                 medical_condition, pat_status, blood_type, allergies, patient_id),
            )
            updated_patients += 1

        # 3. Upsert vitals – update most recent record, or insert if none
        if any([bp_val, heart_rate, temperature, spo2]):
            cur.execute(
                "SELECT id FROM vitals WHERE patient_id = %s ORDER BY last_updated DESC LIMIT 1",
                (patient_id,),
            )
            existing_vitals = cur.fetchone()
            if existing_vitals is None:
                cur.execute(
                    """
                    INSERT INTO vitals (patient_id, bp, heart_rate, temperature, spo2)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (patient_id, bp_val, heart_rate, temperature, spo2),
                )
            else:
                cur.execute(
                    """
                    UPDATE vitals SET
                        bp           = %s,
                        heart_rate   = %s,
                        temperature  = %s,
                        spo2         = %s,
                        last_updated = NOW()
                    WHERE id = %s
                    """,
                    (bp_val, heart_rate, temperature, spo2, existing_vitals[0]),
                )

    conn.commit()
    print(
        f"✓ Patients  – {inserted_patients} new users, "
        f"{updated_patients} records updated, "
        f"{linked_patients} linked to existing staff accounts"
    )

    # ── DYNAMIC SHEETS (newly added Excel tables) ───────────────────────────
    table_columns = get_public_table_columns(cur)
    dynamic_sheets = [s for s in wb.sheetnames if s not in KNOWN_SHEETS]
    if dynamic_sheets:
        print("\n🔎 Processing dynamically-mapped sheets...")
    for sheet_name in dynamic_sheets:
        sheet_records = load_sheet(wb, sheet_name)
        if not sheet_records:
            print(f"↷ Sheet '{sheet_name}' skipped (no data rows)")
            continue
        insert_dynamic_sheet(cur, sheet_name, sheet_records, table_columns)
    conn.commit()

    cur.close()
    conn.close()

    print("\n✅ Excel data fully imported into Neon DB.")
    print(f"\n🔑 Default password for ALL imported accounts: {DEFAULT_PW}")
    print("   (Doctors use their email, e.g. arvindkumar@gmail.com)")


if __name__ == "__main__":
    run()
